# -*- coding: utf-8 -*-

import pydot
import networkx as nx
import angr
# from angrutils import plot_cfg, hook0, set_plot_style
# import bingraphvis
import os
import re
import sys
import subprocess
from natsort import natsorted
import random

import claripy
import multiprocessing
import time
import string
import collections
from multiprocessing import Process, Queue
import io
from tree_sitter import Language, Parser
import csv
import json
from pathlib import Path
import glob
from collections import OrderedDict
from collections import defaultdict
from capstone import *
from networkx.drawing.nx_pydot import write_dot
import argparse
import psutil
from openpyxl import Workbook, load_workbook




def analyze(b, addr, name=None):
    start_state = b.factory.blank_state(addr=addr)
    start_state.stack_push(0x0)
    with hook0(b):
        cfg = b.analyses.CFGEmulated(fail_fast=True, starts=[addr], initial_state=start_state, context_sensitivity_level=2, keep_state=True, call_depth=100, normalize=True)
        #cfg = b.analyses.CFGFast()
    for addr,func in proj.kb.functions.items():
        if func.name in ['main','verify']:
            plot_cfg(cfg, "%s_%s_cfg" % (name, func.name), asminst=True, vexinst=False, func_addr={addr:True}, debug_info=False, remove_imports=True, remove_path_terminator=True)
            plot_cfg(cfg, "%s_%s_cfg" % (name, func.name), asminst=True, vexinst=False, func_addr={addr:True}, debug_info=False, remove_imports=True, remove_path_terminator=True, format="dot")


    plot_cfg(cfg, "%s_cfg_classic" % (name), asminst=True, vexinst=False, debug_info=False, remove_imports=True, remove_path_terminator=True, format="dot")



def getFunAddDict(binary):
    markersAddrFunc = {}
    # print("b = ", b)
    os.system("objdump -S {} > {}.S".format(binary, binary))
    os.system("grep \">$\" {}.S | grep \"<marker_\" > markers-temp-{}.txt".format(binary, binary))
    outFun = subprocess.check_output("cat markers-temp-%s.txt | awk '{print $NF}'" % binary, shell=True)
    # print(outFun.decode())
    outAdd = subprocess.check_output("cat markers-temp-%s.txt | awk '{print $8}'" % binary, shell=True)
    # print(outAdd.decode())
    m = re.findall(r'\d+', outFun.decode())
    n = re.findall(r'[0-9a-zA-Z]+', outAdd.decode())

    j = 0
    for key2 in n:
        markersAddrFunc[key2] = m[j]
        j += 1
    
    return markersAddrFunc


def getCfgDot(binary_file):
    os.system("rm func_*")
    
    try:
        subprocess.check_output("timeout -k 3s 2m bcov -m dump -f \"func_1\" -i %s" % binary_file, shell=True)
    except Exception as e:
        return ""
    # remove some noise assemble code
    os.system("rm func_*rev* func_*dom*")
    os.system("cp func_*.dot {}.dot".format(binary_file))


def getCfgNodeNum(dot_file):
    graphs = pydot.graph_from_dot_file(dot_file)
    graph = graphs[0]
    # convert to nx
    g = nx.drawing.nx_pydot.from_pydot(graph)
    return g.number_of_nodes()


def getAllPaths(dot_file, binary):

    if dot_file == "":
        return

    graphs = pydot.graph_from_dot_file(dot_file)
    # graphs = pydot.graph_from_dot_file("./main.dot")
    # graphs = pydot.graph_from_dot_file("./test-csmith.dot")

    graph = graphs[0]

    # write to png
    # graph.write_png("{}.png".format(dot_file))

    # convert to nx
    g = nx.drawing.nx_pydot.from_pydot(graph)
    print("number of nodes : ", g.number_of_nodes())
    print("number of edges : ", g.number_of_edges())

    # not interesting graph
    if (g.number_of_edges() < 1) or (g.number_of_nodes() < 1):
        print("Not an interesting binary, just skip it ...")
        return

    # g.draw("networkx-test.png")
    i = 1
    start_node = ""
    end_node = ""
    markers = list()
    markers_dict = {}
    markersAddrFunc = getFunAddDict(binary)
    # print("markersAddrFunc : ", markersAddrFunc)
    for node in g.nodes():
        # print(node)
        if node != "" and len(g.nodes()[node]) == 4:
            # print(g.nodes()[node]['label'])
            # print(len(g.nodes()[node]))
            # print(g.nodes()[node])
            # find the marker_start
            str1 = g.nodes()[node]["label"]
            # if (str1.find("start") != -1):
           
            if (str1.find("call") != -1):
                markers.append(node)
                # str1 = str1.replace('\\', '')
                # print("new str1 ", str1)
                # print("match : ", re.findall(r"marker_(.+?)\\", str1))
                temp_nodes = re.findall(r"call (.+?)\\", str1)  # TODO can not match \n for now
                # filter '\\\\n'
                new_temp_nodes = list()
                for temp in temp_nodes:
                    new_temp = temp.replace("\\\n", "")
                    new_temp_nodes.append(new_temp[2:])
                
                if (new_temp_nodes[0] in markersAddrFunc):
                    # print("replacing ++++ ", markersAddrFunc[new_temp_nodes[0]])
                    markers_dict[node] = markersAddrFunc[new_temp_nodes[0]]
                    # set start and node
                    if markersAddrFunc[new_temp_nodes[0]] == "10000":
                        start_node = node
                    if markersAddrFunc[new_temp_nodes[0]] == "19999":
                        end_node = node
                else:
                    markers_dict[node] = -1
                # print("found the marker node ! It is ", node)
        i = i + 1
    
    # no need to know exact paths in CFG, only count the nodes in the CFG and find the difference
    nodes_in_cfg = list()
    for key, value in markers_dict.items():
        if value != -1:
            nodes_in_cfg.append(value)

    return nodes_in_cfg


def checkDiff(list1, list2):
    set1 = set()
    set2 = set()
    for li in list1:
        for l1 in li:
            set1.add(l1)
    for li in list2:
        for l2 in li:
            set2.add(l2)

    n1 = set1
    n2 = set2
    if n1.issubset(n2) is not True or n2.issubset(n1) is not True and len(n1) > 1 and len(n2) > 1:
        print("Find a diff of paths")
        dd1 = n1 - n2
        dd2 = n2 - n1
        diff = set()
        ret = set()
        for d1 in dd1:
            diff.add(d1)
        for d2 in dd2:
            diff.add(d2)
        print("differ : ", diff)
        # check whether they are indeed not existed in .asm
        for d in diff:
            print("Further checking ", d)
            os.system("grep {} test1.asm > /dev/null; echo $? > out1.txt".format(d))
            os.system("grep {} test2.asm > /dev/null; echo $? > out2.txt".format(d))
            grep1 = subprocess.run(['cat', 'out1.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
            grep2 = subprocess.run(['cat', 'out2.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
            if grep1 != grep2:
                ret.add(d)
            else:
                print("May be interesting; skip for now")
    else:
        print("Nothing interesting :-(\n")
        return ret


# interesting means markers are union(all path) - intersect(all path)
def storeAllPaths(binary):
    all_path_dict = {}
    for b in binary:
        getCfgDot(b)
        # print("file exist? ", os.path.isfile(b))
        if os.path.isfile("{}.dot".format(b)) is True:  # .dot is generated
            paths = getAllPaths("{}.dot".format(b), b)
            all_path_dict[b] = paths
        else:
            break
    return all_path_dict


# interesting means markers are union(all path) - intersect(all path)
def getInterestingMarkerSetV1(d):
    markers = set()
    new_dict = {}
    binary = []
    for key, value in d.items():
        # print("key : ", key)
        # print("value : ", value)
        binary.append(key)
        # transfer list to set first
        m = set()
        if value is None:
            return None
        for ss in value:
            m.add(ss)
        new_dict[key] = m
    # print("new_dict : ", new_dict)
    b_len = len(new_dict)
    if b_len < 2:
        return
    # print("b_len : ", b_len)
    marker_intersect = list()
    for i in range(b_len):
        # print(markers_all[binary[i]])
        # print(markers_all[binary[i+1]])
        marker_intersect = new_dict[binary[i]] & new_dict[binary[i+1]]
        if i+1 == b_len - 1:
            break
    # get interesting markers
    for i in range(b_len):
        markers.update(new_dict[binary[i]] - marker_intersect)
    return markers


def remove_common(a, b):
    for i in a[:]:
        if i in b:
            a.remove(i)
            b.remove(i)
    return a, b


def check_markers_in_cfg(b1, b2, marker):
    os.system("objdump -S {} > {}.S".format(b1, b1))
    os.system("grep \">$\" {}.S | grep \"<marker_{}\" > markers-{}.txt".format(b1, marker, b1))
    os.system("objdump -S {} > {}.S".format(b2, b2))
    os.system("grep \">$\" {}.S | grep \"<marker_{}\" > markers-{}.txt".format(b2, marker, b2))
    grep1 = subprocess.run(['cat', 'markers-test1.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
    grep2 = subprocess.run(['cat', 'markers-test2.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
    os.system("cat markers-test1.txt")
    os.system("cat markers-test2.txt")
    if grep1 == grep2:
        return True
    else:
        return False


# interesting means markers are set per path
def getInterestingMarkerSetV2(d):
    markers = set()
    new_dict = {}
    # first remove the same path in d
    if d["test1"] is None or d['test2'] is None:
        return None

    print("test ---------------")
    print(type(d['test1']))
    a, b = remove_common(d['test1'], d['test2'])
    print("a : ", a)
    print("b : ", b)
    new_dict = {}
    new_dict["test1"] = a
    new_dict["test2"] = b
    len_test1 = len(new_dict['test1'])
    len_test2 = len(new_dict['test2'])
    print("len_test1 : ", len_test1)
    print("len_test2 : ", len_test2)
    if len_test1 == 0 and len_test2 == 0:   # no interesting paths
        return None
    if len_test1 != 0 and len_test2 != 0:
        print("test ---------------")
        # path to path comparison
        for t1 in new_dict["test1"]:
            for t2 in new_dict["test2"]:
                list_temp = set(t1).union(set(t2)) - set(t1).intersection(set(t2))
                for m in list_temp:
                    markers.add(m)
    else:
        if len_test1 == 0:  # test1 is empty
            print("test1 is empty; to be added")  # TODO return none for now
            return None
        else:  # test2 is empty
            print("test2 is empty")  # TODO return none for now
            return None
    return markers


# interesting means markers in a cycle
def getInterestingMarkerSetV3(d):
    markers = set()
    path_set1 = set()
    path_set2 = set()
    paths1 = d["test1"]
    paths2 = d['test2']
    if paths1 is not None:
        for path1 in paths1:
            for p1 in path1:
                path_set1.add(p1)
    if paths2 is not None:
        for path2 in paths2:
            for p2 in path2:
                path_set2.add(p2)
    interesting_marker1 = path_set1.union(path_set2) - path_set1
    interesting_marker2 = path_set1.union(path_set2) - path_set2
    markers = interesting_marker1 | interesting_marker2  # add two sets to be one
    return markers


# dead code version
def getInterestingMarkerSetV4(d):
    markers = dict()
    path_set1 = set()
    path_set2 = set()
    paths1 = d["test1"]
    paths2 = d['test2']
    if paths1 is not None:
        for path1 in paths1:
            # for p1 in path1:
            path_set1.add(path1)
    if paths2 is not None:
        for path2 in paths2:
            # for p2 in path2:
            path_set2.add(path2)
    intersect = path_set1.intersection(path_set2)
    print("intersect : ", intersect)
    interesting_marker1 = path_set1 - intersect
    interesting_marker2 = path_set2 - intersect
    markers["test1"] = interesting_marker1
    markers["test2"] = interesting_marker2
    return markers


def filter_intersection(orig_intersect, dot_file, binary):
    # basic idea: get the original marker list, and cut the markers in the same path
    # step 1: preparing the checking, i.e., dictionary that records the mapping between node and markers
    if dot_file == "":
        return

    graphs = pydot.graph_from_dot_file(dot_file)

    graph = graphs[0]

    # convert to nx
    g = nx.drawing.nx_pydot.from_pydot(graph)
    i = 1
    start_node = ""
    end_node = ""
    markers = list()
    markers_dict = {}
    markersAddrFunc = getFunAddDict(binary)
    # print("markersAddrFunc : ", markersAddrFunc)
    for node in g.nodes():
        # print(node)
        if node != "" and len(g.nodes()[node]) == 4:
            str1 = g.nodes()[node]["label"]
            # if (str1.find("start") != -1):
            if (str1.find("call") != -1):
                markers.append(node)
                temp_nodes = re.findall(r"call (.+?)\\", str1)  # TODO can not match \n for now
                new_temp_nodes = list()
                for temp in temp_nodes:
                    new_temp = temp.replace("\\\n", "")
                    new_temp_nodes.append(new_temp[2:])
                if (new_temp_nodes[0] in markersAddrFunc):
                    # print("replacing ++++ ", markersAddrFunc[new_temp_nodes[0]])
                    markers_dict[node] = markersAddrFunc[new_temp_nodes[0]]
                    # set start and node
                    if markersAddrFunc[new_temp_nodes[0]] == "10000":
                        start_node = node
                    if markersAddrFunc[new_temp_nodes[0]] == "19999":
                        end_node = node
                else:
                    markers_dict[node] = -1
        i = i + 1
    # print the markers_dict
    #print("markers_dict : ", markers_dict)
    # no need to know exact paths in CFG, only count the nodes in the CFG and find the difference
    nodes_in_cfg = list()
    for key, value in markers_dict.items():
        if value != -1:
            nodes_in_cfg.append(value)
    # print("test filter_intersection : ", nodes_in_cfg)

    # check and whether every node is in the CFG
    orig_intersect_new = list()
    for m in orig_intersect:
        if m in nodes_in_cfg:
            orig_intersect_new.append(m)

    if len(orig_intersect_new) == 0:
        print("Do not exist in this CFG ...")
        return None

    # directly return if there is only one node in the intersection
    if len(orig_intersect_new) < 2:
        return orig_intersect_new
    # check whether some of the nodes are in the same path
    # order first
    ordered_ori_list = natsorted(orig_intersect_new)
    print("ordered_ori_list : ", ordered_ori_list)
    # every time check two nodes TODO better solution?
    final_list = list()
    pindex = 0
    sources_temp = list()
    
    nindex = 0
    while (pindex + 1) < len(orig_intersect):
        # print("while?")
        # find the node in the dictory
        if len(sources_temp) < 2:
            # print("< 2 : nindex", nindex)
            # print("< 2 : pindex", pindex)
            for key, value in markers_dict.items():
                if nindex + 1 + pindex >= len(ordered_ori_list):
                    return final_list
                if value == ordered_ori_list[nindex + pindex]:
                    sources_temp.append(key)
                    # print("value 1: ", value)
                if value == ordered_ori_list[nindex + 1 + pindex]:
                    sources_temp.append(key)
                    # print("value 2: ", value)
            nindex += 2
        else:
            for key, value in markers_dict.items():
                if nindex + 1 + pindex >= len(ordered_ori_list):
                    return final_list
                if value == ordered_ori_list[nindex + pindex]:
                    sources_temp.append(key)
                    # print("value ?: ", value)
            # pindex += 1
        if nx.is_simple_path(g, sources_temp) is True and nindex < len(ordered_ori_list):
            print("in the same path")
            # check whether pindex is pointing to the last two elements
            if pindex == len(ordered_ori_list) - 2:
                if markers_dict[sources_temp[0]] not in final_list:
                    final_list.append(markers_dict[sources_temp[0]])
                pindex += 1
        else:
            print("not in the same path")
            if markers_dict[sources_temp[0]] not in final_list:
                final_list.append(markers_dict[sources_temp[0]])
            pindex += len(sources_temp) - 1
            # same_path_count += 1
            sources_temp.clear()  # clear the list
            nindex = 0

    return final_list


# Define directories and file paths
WORKSPACE_ROOT = os.path.dirname(os.path.abspath(__file__))
TESTCASES_DIR = os.path.join(WORKSPACE_ROOT, "testcases")
RESULTS_DIR = os.path.join(WORKSPACE_ROOT, "results")

# Ensure the results directory exists
os.makedirs(RESULTS_DIR, exist_ok=True)

def get_testcase_files():
    """Get all .c files in the testcases directory"""
    return sorted([f for f in glob.glob(os.path.join(TESTCASES_DIR, "testcase_*.c"))])

def create_codeql_db(testcase_path, db_path):
    """Create a CodeQL database for a single test case"""
    testcase_filename = os.path.basename(testcase_path)
    #print(f"\n[+] Creating CodeQL database for {testcase_filename}...")

    # Build the database creation command
    cmd = [
        "codeql", "database", "create",
        "--language=cpp",
        f"--command=gcc -c {testcase_path} -I /usr/include/csmith/",
        "--overwrite",
        db_path,
        f"--source-root={os.path.dirname(testcase_path)}"
    ]

    try:
        # result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, check=True)
        #print(f"[+] Successfully created database: {db_path}")
        return True
    except subprocess.CalledProcessError as e:
        print(f"[!] Failed to create database: {e}")
        print(f"Error output: {e.stderr}")
        return False

def run_codeql_query(db_path, query_path, csv_output_path):
    """Run the CodeQL query on the database and save directly to CSV format"""
    #print(f"[+] Running query on {os.path.basename(db_path)}...")

    # Build the query command, using database analyze to directly output CSV
    cmd = [
        "codeql", "database", "analyze",
        "--format=csv",
        f"--output={csv_output_path}",
        db_path,
        query_path
    ]

    try:
        # result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, check=True)
        #print(f"[+] Query successful, results saved to: {csv_output_path}")
        return True
    except subprocess.CalledProcessError as e:
        print(f"[!] Query failed: {e}")
        print(f"Error output: {e.stderr}")
        return False

def parse_csv_to_dict(csv_file_path):
    """
    Parse the CSV file and save the results to a dictionary.
    :param csv_file_path: CSV file path
    :return: A dictionary where the key is the global variable name and the value is a list of attributes including line numbers.
    """
    result_dict = {}
    try:
        with open(csv_file_path, mode='r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                if len(row) < 5:
                    continue
                # Extract nested JSON data
                json_data = row[3]
                try:
                    # Fix JSON data format issues, replacing ";" with ","
                    fixed_json_data = json_data.replace(';', ',')
                    # j = f"[{fixed_json_data.replace('}\n{', '},{')}]"
                    replaced_data = fixed_json_data.replace('}\n{', '},{')
                    j = f"[{replaced_data}]"
                    parsed_data = json.loads(j)
                    for entry in parsed_data:
                        global_var_name = entry["globalVarName"]
                        location = entry["location"]
                        # Extract source line from the file
                        src_line = ""
                        try:
                            with open(os.path.join(TESTCASES_DIR, row[4][1:]), 'r') as f:
                                lines = f.readlines()
                                line_num = int(location)
                                if 0 < line_num <= len(lines):
                                    src_line = lines[line_num - 1].strip()
                        except Exception as e:
                            print(f"[!] Failed to read source line: {e}")
                            import traceback; traceback.print_exc()

                        attributes = {
                            "type": entry["globalVarType"],
                            "location": location,
                            "reason": entry["reason"],
                            "sinkVarName": entry["sinkVarName"],
                            "src": src_line
                        }
                        if global_var_name not in result_dict:
                            result_dict[global_var_name] = []
                        result_dict[global_var_name].append(attributes)
                except json.JSONDecodeError as e:
                    print(f"[!] JSON parsing failed: {e}")
    except Exception as e:
        print(f"[!] Failed to parse CSV file: {e}")
    return result_dict


def process_testcase_for_var_selection(testcase_path):
    """Optimized Tree-Sitter-based analysis with slicing and early stop"""

    import os
    import json
    from tree_sitter import Language, Parser

    LANG_SO_PATH = 'build/my-languages.so'
    if not os.path.exists(LANG_SO_PATH):
        Language.build_library(LANG_SO_PATH, ['/home/xxx/disk-dut/research/github/tree-sitter-c'])

    C_LANGUAGE = Language(LANG_SO_PATH, 'c')

    testcase_name = os.path.basename(testcase_path).split('.')[0]
    json_results_path = os.path.join(RESULTS_DIR, f"{testcase_name}_results.json")

    print(f"========== Processing {testcase_name} with Data-Flow Analysis ==========")

    parser = Parser()
    parser.set_language(C_LANGUAGE)

    global count_var_org, count_var_reduced
    with open(testcase_path, 'r', encoding='utf-8') as f:
        source_code = f.read()

    tree = parser.parse(source_code.encode('utf-8'))
    root_node = tree.root_node

    # Step 1: Find global variables (with integer types, starting with g_)
    global_vars = {}

    def find_globals(node):
        if node.type == 'declaration' and 'extern' not in node.text.decode():
            var_decl = node.child_by_field_name('declarator')
            var_type = node.child_by_field_name('type')
            if var_decl and var_type:
                var_name = var_decl.text.decode().split('=')[0].strip()
                type_text = var_type.text.decode()
                if var_name.startswith('g_') and any(k in type_text for k in ('int', 'uint', 'long', 'short', 'int8', 'int16', 'int32', 'int64', 'uint8', 'uint16', 'uint32', 'uint64')):
                    global_vars[var_name] = type_text
        for child in node.children:
            find_globals(child)

    find_globals(root_node)

    print("[-] Number of global variables (before OPT) : ", len(global_vars))
    count_var_org += len(global_vars)
    # Step 2: Collect all condition expressions
    condition_nodes = []

    def collect_conditions(node):
        if node.type in ('if_statement', 'while_statement', 'for_statement', 'do_statement', 'conditional_expression'):
            cond_child = node.child_by_field_name('condition')
            if cond_child:
                condition_nodes.append((node, cond_child))
        for child in node.children:
            collect_conditions(child)

    collect_conditions(root_node)

    # Step 3: Build assignment slices {var_name: set of source variables}
    assignments = {}

    def collect_assignments(node):
        if node.type == 'assignment_expression':
            left = node.child_by_field_name('left')
            right = node.child_by_field_name('right')
            if left and right and left.type == 'identifier':
                target = left.text.decode()
                sources = set()

                def collect_rhs_vars(rhs_node):
                    if rhs_node.type == 'identifier':
                        sources.add(rhs_node.text.decode())
                    for child in rhs_node.children:
                        collect_rhs_vars(child)

                collect_rhs_vars(right)
                if target not in assignments:
                    assignments[target] = set()
                assignments[target].update(sources)
        for child in node.children:
            collect_assignments(child)

    collect_assignments(root_node)

    # Step 4: Slicing to find indirect global variable usage
    def slice_back(var, visited):
        if var in visited:
            return set()
        visited.add(var)
        if var in global_vars:
            return {var}
        result = set()
        for src in assignments.get(var, []):
            result.update(slice_back(src, visited))
        return result

    # Step 5: Analyze each condition
    found_globals = set()
    results = {}

    def get_vars_in_expr(node, collected):
        if node.type == 'identifier':
            collected.add(node.text.decode())
        for child in node.children:
            get_vars_in_expr(child, collected)

    for control_node, condition_expr in condition_nodes:
        vars_in_cond = set()
        get_vars_in_expr(condition_expr, vars_in_cond)

        cond_line = str(condition_expr.start_point[0] + 1)

        for var in vars_in_cond:
            if var in found_globals:
                continue

            if var in global_vars:
                results[var] = [{
                    "type": global_vars[var],
                    "location": cond_line,
                    "reason": "direct_usage",
                    "sinkVarName": var
                }]
                found_globals.add(var)
                continue

            slice_result = slice_back(var, set())
            for gvar in slice_result:
                if gvar not in found_globals:
                    results[gvar] = [{
                        "type": global_vars[gvar],
                        "location": cond_line,
                        "reason": "indirect_usage",
                        "sinkVarName": var
                    }]
                    found_globals.add(gvar)

        if len(found_globals) == len(global_vars):
            break

    print("[+] reduced number of global variables (after OPT) : ", len(found_globals))
    count_var_reduced += len(found_globals)
    with open(json_results_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=4, ensure_ascii=False)
    # print(f"[+] JSON results saved to: {json_results_path}")

    return True


compare = lambda x, y: collections.Counter(x) == collections.Counter(y)


def get_hex(text):
    valid_set = '0123456789abcdefABCDEF'
    sym_set = '+-x'
    res = ""
    
    for i in text:
        if valid_set.find(i) != -1:
            # if valid_set.find(i) != -1 and (i == 'x' and (index == 1 or index == 2)):
            # and text[len(text) - 1] != '+' and text[len(text) - 1] != '-' and text[len(text) - 1] != 'x':
            res += i
        elif index == 1 and (text[0] != '-' or text[0] != '+') and text[index - 1] == '0' and text[index] == 'x':  # 0x
            res += i
        elif index == 2 and (text[0] == '-' or text[0] == '+') and text[index - 1] == '0' and text[index] == 'x':
            res += i
        elif index == 0 and (text[0] == '-' or text[0] == '+'):
            res += i
        else:
            break
        index += 1
    # print("before handling 0x : ", res)
    new_res = ""
    if len(res) == 0:
        new_res = "0"
        return new_res
    # if res.find('0x') != -1 and res[0] != '-' and res[0] != '+':
    #    return res
    if res.find('0x') != -1:  # it has 0x
        # new_res = res[0] + "0x" + res[1:]
        if res == "-0x" or res == "+0x" or res == "0x":
            return "0"
        else:
            new_res == res
    # else:
    #    new_res = "0x" + res
    #    return new_res
    # print("res ", res.find('0x'), res[0], new_res)
    if res.find('0x') == -1 and res[0] == '-':  # doesn't have 0x
        new_res = res[0] + '0x' + res[1:]
    elif res.find('0x') == -1 and res[0] == '+':  # doesn't have 0x
        new_res = res[0] + '0x' + res[1:]
    elif res.find('0x') == -1 and res[0] != '-' and res[0] != '+':  # doesn't have 0x
        # return new_res
        new_res = '0x' + res
    else:
        new_res = res
        # return new_res
    # cut final "+-x"
    # print("before cutting: ", new_res)
    c = new_res[len(new_res) - 1]
    while c == '+' or c == '-' or c == 'x':
        new_res = new_res[:len(new_res) - 1]
        c = new_res[len(new_res) - 1]
    return new_res


def remove_dup(ss):
    return list(set(ss))


def remove_dup_con(con):
    ret = []
    ret_temp = []
    temp = set()
    for cc in con:
        for c in cc:
            temp.add(c)
    if temp not in ret_temp:
        ret.append(cc)
    return ret


def is_the_same(str_set1, str_set2):
    # covert the set to string list
    # if len(str_set1) == 0 or len(str_set2) == 0:  # skip zero cases? TODO
    #    return True
    str_list1 = []
    str_list2 = []
    for s1 in str_set1:
        s1_list = s1.split()
        str_list1.extend(s1_list)
    for s2 in str_set2:
        s2_list = s2.split()
        str_list2.extend(s2_list)
    # print(str_list1)
    # print(str_list2)
    if len(str_list1) == len(str_list2):  # have the same length
        str_list1_set = set(str_list1)
        str_list2_set = set(str_list2)
        print("str_list1_set : ", str_list1_set)
        print("str_list2_set : ", str_list2_set)
        if str_list1_set.issubset(str_list2_set) and str_list2_set.issubset(str_list1_set):
            return True
        else:
            return False
    else:
        return False


def checking_binary_diff(binary1, binary2, angr_set1, angr_set2):
    os.system("objdump -d b-marker1 | grep marker | grep \"call\|jmp\|callq\|jne\" | awk \'{print $9}\' > out1.txt")
    os.system("objdump -d b-marker2 | grep marker | grep \"call\|jmp\|callq\|jne\" | awk \'{print $9}\'> out2.txt")

    list1 = []
    with open("out1.txt") as file_in:
        for line in file_in:
            line = re.findall(r"\d+\.?\d*", line)
            list1.append(line)
    set1 = set()
    for ss in list1:
        for s in ss:
            set1.add(s)

    list2 = []
    with open("out2.txt") as file_in:
        for line in file_in:
            line = re.findall(r"\d+\.?\d*", line)
            list2.append(line)

    set2 = set()
    for ss in list2:
        for s in ss:
            set2.add(s)
    # n1 = set(list1)
    # n2 = set(list2)
    # print("before comparing set1: ", set1)
    # print("before comparing set2: ", set2)

    n1 = set1
    n2 = set2
    if n1.issubset(n2) is not True or n2.issubset(n1) is not True and len(n1) > 1 and len(n2) > 1:
        print("Find a diff of binary")
        dd1 = n1 - n2
        dd2 = n2 - n1
        diff = set()
        diff_se = set()
        for d1 in dd1:
            diff.add(d1)
        for d2 in dd2:
            diff.add(d2)
        for s1 in angr_set1:
            diff_se.add(s1)
        for s2 in angr_set2:
            diff_se.add(s2)
        if diff.issubset(diff_se) is True or diff_se.issubset(diff) is True:
            return True
        else:
            return False
    else:
        print("Nothing interesting :-(\n")
        return False

def checking_markers(binary1, binary2):
    os.system("objdump -d test1 | grep marker | grep \"call\\|jmp\\|callq\\|jne\" | awk \'{print $9}\' > out1.txt")
    os.system("objdump -d test2 | grep marker | grep \"call\\|jmp\\|callq\\|jne\" | awk \'{print $9}\'> out2.txt")

    list1 = []
    with open("out1.txt") as file_in:
        for line in file_in:
            line = re.findall(r"\d+\.?\d*", line)
            list1.append(line)
    set1 = set()
    for ss in list1:
        # for s in ss:
        set1.add(ss)

    list2 = []
    with open("out2.txt") as file_in:
        for line in file_in:
            line = re.findall(r"\d+\.?\d*", line)
            list2.append(line)

    set2 = set()
    for ss in list2:
        # for s in ss:
        set2.add(ss)
    # n1 = set(list1)
    # n2 = set(list2)
    # print("before comparing set1: ", set1)
    # print("before comparing set2: ", set2)
    n1 = set1
    n2 = set2
    if n1.issubset(n2) is not True or n2.issubset(n1) is not True and len(n1) > 1 and len(n2) > 1:
        # print("Find a diff of binary")
        dd1 = n1 - n2
        dd2 = n2 - n1
        diff = set()
        diff_se = set()
        for d1 in dd1:
            diff.add(d1)
        for d2 in dd2:
            diff.add(d2)
        return True, diff
    else:
        # print("Nothing interesting :-(\n")
        return False, {}


def is_diff_set(set1, set2):
    n1 = set1
    n2 = set2
    if n1.issubset(n2) is not True or n2.issubset(n1) is not True and len(n1) > 1 and len(n2) > 1:
        # print("Find a diff of binary")
        dd1 = n1 - n2
        dd2 = n2 - n1
        diff = set()
        diff_se = set()
        for d1 in dd1:
            diff.add(d1)
        for d2 in dd2:
            diff.add(d2)
        return True, diff
    else:
        # print("Nothing interesting :-(\n")
        return False, {}


def test_native():
    # record name
    name_list = []
    with open("names.txt") as file_in:
        for line in file_in:
            name_list.append(line.rstrip('\n'))
    print("name_list : ", name_list, len(name_list))

    concrete_value = []
    with open("temp.txt") as file_in:
        for line in file_in:
            concrete_value.append(line.rstrip('\n'))
    # cut and split
    value_list = concrete_value[0].split(" ")
    value_list = value_list[:-1]
    print("value_list : ", value_list, len(value_list))
    new_assign = []
    # cut the last ''
    if len(name_list) != len(value_list):
        print("Error: the number of name and concrete value is not matched ...")
        exit(1)
    for i in range(len(name_list)):
        s = name_list[i] + " = " + value_list[i] + " ;"
        new_assign.append(s)
    # print(new_assign)
    orig_assign = []
    for name in name_list:
        out = subprocess.check_output("grep \"%s = strto\" testcase_1.c" % name, shell=True)
        # print(out.decode())
        orig_assign.append(out.decode()[4:-2])
    # print(orig_assign)
    # replace
    os.system("cp testcase_1.c test-csmith.c")
    for j in range(len(name_list)):
        # print("orig_assign[i] : ", orig_assign[j])
        # print("new_assign[i] : ", new_assign[j])
        os.system("sed -i \'s/{} = strto.*/{}/g\' test-csmith.c".format(name_list[j], new_assign[j]))
        # os.system("cat testcase_1.c")
    print("Replace variables done and string native running and checking ...")
    # conduct compiler testing
    os.system("./compiler_test.pl 1 compiler_test.in")


def getAddrMarkerDict(binary):
    d = {}
    if binary == 'test1':
        os.system("objdump -d test1 | grep marker | grep -E \'\\b0000\' | awk \'{print $1}\' > out-addr.txt")
        os.system("objdump -d test1 | grep marker | grep -E \'\\b0000\' | awk \'{print $2}\'> out-marker.txt")
    if binary == 'test2':
        os.system("objdump -d test2 | grep marker | grep -E \'\\b0000\' | awk \'{print $1}\' > out-addr.txt")
        os.system("objdump -d test2 | grep marker | grep -E \'\\b0000\' | awk \'{print $2}\'> out-marker.txt")
    addr_list = []
    with open("out-addr.txt") as file_in:
        for line in file_in:
            res = line.rstrip('\n')
            res = res.replace('0000000000', '')
            res = '0x' + res
            addr_list.append(res)
    # print("addr_list : ", addr_list)
    marker_list = []
    with open("out-marker.txt") as file_in:
        for line in file_in:
            res = line.rstrip(':\n')
            res = res.replace('<', '')
            res = res.replace('>', '')
            marker_list.append(res)
    # print("marker_list : ", marker_list)
    if len(addr_list) != len(marker_list):
        print("Fatal ERROR: address and marker are not matched!!!")
        exit(0)
    for i in range(len(addr_list)):
        d[addr_list[i]] = marker_list[i]
    return d


def getInterestingMarkerSetV5():
    # os.system("objdump -d test1 | grep \"<marker_*\" | awk \'{print $2}\' | grep marker > out1.txt")
    # os.system("objdump -d test2 | grep \"<marker_*\" | awk \'{print $2}\' | grep marker > out2.txt")
    os.system("objdump -d test1 | grep marker | grep \"call\\|jmp\" | awk \'{print $9}\' > out1.txt")
    os.system("objdump -d test1 | grep marker | grep \"jmp\" | awk \'{print $6}\' >> out1.txt")
    os.system("objdump -d test1 | grep marker | grep \"jne\" | awk \'{print $10}\' >> out1.txt")
    os.system("objdump -d test1 | grep marker | grep \"je\" | awk \'{print $10}\' >> out1.txt")
    os.system("objdump -d test1 | grep marker | grep \"js\" | awk \'{print $10}\' >> out1.txt")
    os.system("objdump -d test1 | grep marker | grep \"j*\" | awk \'{print $10}\' >> out1.txt")
    os.system("objdump -d test2 | grep marker | grep \"call\\|jmp\" | awk \'{print $9}\'> out2.txt")
    os.system("objdump -d test2 | grep marker | grep \"jmp\" | awk \'{print $6}\'>> out2.txt")
    os.system("objdump -d test2 | grep marker | grep \"jne\" | awk \'{print $10}\'>> out2.txt")
    os.system("objdump -d test2 | grep marker | grep \"je\" | awk \'{print $10}\'>> out2.txt")
    os.system("objdump -d test2 | grep marker | grep \"js\" | awk \'{print $10}\'>> out2.txt")
    os.system("objdump -d test2 | grep marker | grep \"j*\" | awk \'{print $10}\' >> out2.txt")

    list1 = []
    with open("out1.txt") as file_in:
        for line in file_in:
            # result = re.search('%s(.*)%s' % ("_", "_"), line).group(1)
            # list1.append(result)
            # print("result :", result)
            line = re.findall(r"\d+\.?\d*", line)
            list1.append(line)
    set1 = set()
    for ss in list1:
        for s in ss:
            set1.add(s)

    list2 = []
    with open("out2.txt") as file_in:
        for line in file_in:
            line = re.findall(r"\d+\.?\d*", line)
            list2.append(line)
            # result = re.search('%s(.*)%s' % ("_", "_"), line).group(1)
            # list2.append(result)

    set2 = set()
    for ss in list2:
        for s in ss:
            set2.add(s)
    # n1 = set(list1)
    # n2 = set(list2)
    print("before comparing set1: ", set1)
    print("before comparing set2: ", set2)
    markers = dict()
    path_set1 = set()
    path_set2 = set()
    paths1 = set1
    paths2 = set2
    if paths1 is not None:
        for path1 in paths1:
            # for p1 in path1:
            path_set1.add(path1)
    if paths2 is not None:
        for path2 in paths2:
            # for p2 in path2:
            path_set2.add(path2)
    intersect = path_set1.intersection(path_set2)
    print("intersect : ", intersect)
    interesting_marker1 = path_set1 - intersect
    interesting_marker2 = path_set2 - intersect
    markers["test1"] = interesting_marker1
    markers["test2"] = interesting_marker2
    return markers


def targeted_se(target_binary, target_marker):
    # Main process
    # 1. Get the sequences of type and size of each variable from `csmith-se`

    # 2. Make those variables symbolic and control their range; TODO range can be done later

    # 3. Conduct binary symbolic execution by angr

    # 4. Compare and record results; path explored and record concrete values of symbolic variables

    # 5. Get each set of concrete values and re-run with test program generated by `csmith-run`

    #
    # Step 1
    #

    target_binary1 = target_binary

    d = getAddrMarkerDict(target_binary1)


    results = subprocess.run(['tail', '-1', 'testcase_1.c'], stdout=subprocess.PIPE).stdout.decode("utf-8")

    print(results[2:-1])
    print(len(results[2:-1]))

    len_results = len(results[2:-1])

    if len_results % 2 != 0:
        print("Error: check return from source code testcase_1.c !")
        exit(1)

    var_sym = []
    var_size = []
    dict_sym = {}

    for i in range(0, len_results, 2):
        var_sym.append(results[2+i])
        var_size.append(results[2+i+1])

    # print(var_sym)
    # print("number of symbolic variable: ", var_size)

    for i in range(len(var_sym)):
        dict_sym['arg{}'.format(i)] = [var_sym[i], var_size[i]]
    #
    # Step 2
    #

    # make symbolic for each variable
    keys = []
    for key, value in dict_sym.items():
        # print(key)
        locals()[key] = claripy.BVS(key, 8*int(value[1])*2)
        keys.append(locals()[key])

    # for i in range(len(var_sym)):
    #     keys += "arg" + str(i) + ", "
    # print(keys)

    # os.system("{} -w -{} -o b-{}-{}".format(cmp1, o, cmp1, o))
    # os.system("{} -w -{} -o b-{}-{}".format(cmp2, o, cmp2, o))


    # check markers first
    interesting_markers = target_marker
    print("find_str : ", interesting_markers)

    # here find_bytes should be a sequence

    #find_str_dig = find_str_dig[1:len(find_str_dig)-1]
    find_str_dig = list(interesting_markers)  # no need to cut start/end
    print("find_str_dig seq : ", find_str_dig)
    find_seq = ""
    for s in find_str_dig:
        find_seq += s
    find_seq = find_seq.replace(",", "")
    find_seq = find_seq.replace("[", "")
    find_seq = find_seq.replace("]", "")
    find_seq = 'b' + find_seq + 'b'
    print("find_seq: ", find_seq)

    find_bytes = bytes(find_seq, "utf-8")


    # for target_binary1
    p1 = angr.Project(target_binary1, load_options={'auto_load_libs': False})
    state1 = p1.factory.entry_state(args=[target_binary1, *keys], add_options={angr.options.ZERO_FILL_UNCONSTRAINED_MEMORY, angr.options.ZERO_FILL_UNCONSTRAINED_REGISTERS})

    # add more constraints for signed/unsigned issue
    for i in range(len(var_sym)):
        # print(keys[i])
        if var_sym[i] == 0 and var_type[i] == 1:  # int8_t
            state1.solver.add(-128 <= keys[i] <= 127)
        if var_sym[i] == 0 and var_type[i] == 2:  # int16_t
            state1.solver.add(-32768 <= keys[i] <= 32767)
        if var_sym[i] == 0 and var_type[i] == 4:  # int32_t
            state1.solver.add(-2147483648 <= keys[i] <= 2147483647)
        if var_sym[i] == 0 and var_type[i] == 8:  # long or long long or int64_t
            state1.solver.add(-9223372036854775808 <= keys[i] <= 9223372036854775807)
            # state1.solver.add(0 < keys[i] < 9223372036854775807)
        if var_sym[i] == 1 and var_type[i] == 1:  # uint8_t
            state1.solver.add(0 <= keys[i] <= 255)
        if var_sym[i] == 1 and var_type[i] == 2:  # uint16_t
            state1.solver.add(0 <= keys[i] <= 65535)
        if var_sym[i] == 1 and var_type[i] == 4:  # uint32_t
            state1.solver.add(0 <= keys[i] <= 4294967295)
        if var_sym[i] == 1 and var_type[i] == 8:  # uint64_t
            state1.solver.add(0 <= keys[i] <= 18446744073709551615)


    sm1 = p1.factory.simulation_manager(state1)
    sm1.use_technique(angr.exploration_techniques.local_loop_seer.LocalLoopSeer(bound=10000))
    print("Start to run targeted symbolic execution ....")
    # sm1.run()
    print("find_bytes = ", find_bytes)
    sm1.explore(find=lambda s: find_bytes in s.posix.dumps(1)) # find the target only

    print("Run targeted symbolic execution done ...")
    # sm1.explore(find=lambda s: find_bytes in s.posix.dumps(1))
    # print(sm1.deadended)
    print(sm1)
    # str11 = sm1.deadended[0].solver.constraints
    # print(str11)
    testcase1 = []
    constraints1 = []
    output1 = []
    output11 = []
    raw_results1 = []
    for i in range(len(sm1.found)):
        o1 = sm1.found[i].posix.dumps(1).decode('utf-8')
        output1.append(o1)
        temp_list = []
        raw = []
        for arg in keys:
            b = sm1.found[i].solver.eval(arg, cast_to=bytes).rstrip(b'\x00')
            b_strstr = str(b).split("\\x")[0]
            if b_strstr.find('\'') != -1:
                b_str = str(b).split("\\x")[0].split('\'')[1]
            else:
                b_str = ""
            raw.append(b_str)
            temp_list.append(get_hex(b_str))
        # print(temp_list)
        testcase1.append(temp_list)
        raw_results1.append(raw)

    #print("raw_results1 : ", raw_results1)
    print("output1 : ", output1)
    # print("output11 : ", output11)

    # compare results and perform recording
    file_temp = open("temp.txt", "w")


    for tt in testcase1:
        for t in tt:
            file_temp.write(t + " ")

    file_temp.write("\n")
    file_temp.close()
    # os.system("cat temp.txt")
    # TODO extract markers from output1 and compare with the interesting ones
    print("interesting_markers : ", interesting_markers)
    markers = set()

    # get all the interesting markers from CFG
    all_path_str1 = ""
    for p1 in output1:
        all_path_str1 = all_path_str1 + p1 + " "

    path_numbers1 = re.findall(r'\d+', all_path_str1)
    print("path_numbers : ", path_numbers1)
    set_markers = interesting_markers
    set_paths1 = path_numbers1

    print("set_markers : ", set_markers)
    print("set_paths1 : ", set_paths1)


    if set_markers in set_paths1:  # interesting
    # if ret1[0] is False and len(set_paths1) != 0:  # interesting
        print("Found a difference HERE!")
        global count_found
        count_found += 1
        # native run
        # exit(1)
        num = int(random.random()*100000000)
        os.system("cp testcase_1.c wrong-{}-candidate.c".format(num))
        os.system("cat temp.txt >> wrong-{}-candidate.c".format(num))
        test_native()
        os.system("cp test-csmith.c wrong-{}-candidate-replaced.c".format(num))

        return True
    else:
        print("No semantic divergence Found -:(\n")
        return False
        # os.system("rm b-mar*")


def wrapper(b, inter, q):
    result = targeted_se(b, inter)
    q.put(result)
    print("Return value:", result)


def instrument_c_file(input_file: str, output_file: str):
    # === Setup Tree-sitter ===
    Language.build_library(
        'build/my-languages.so',
        ['/home/xxx/disk-dut/research/github/tree-sitter-c']
    )

    C_LANGUAGE = Language('build/my-languages.so', 'c')
    parser = Parser()
    parser.set_language(C_LANGUAGE)

    with open(input_file, "rb") as f:
        code = f.read()

    tree = parser.parse(code)

    # === Check if a node is inside main() ===
    def is_inside_main(node):
        while node:
            if node.type == "function_definition":
                for child in node.children:
                    if child.type == "declarator":
                        ident = child.child_by_field_name("declarator")
                        if ident and code[ident.start_byte:ident.end_byte] == b"main":
                            return True
            node = node.parent
        return False

    # === Collect insertion points (skip main) ===
    def collect_marker_insertions(node, marker_positions):
        if node.type == 'compound_statement' and not is_inside_main(node):
            children = node.children
            insert_pos = None

            for i in reversed(range(len(children))):
                child = children[i]
                if child.type in ['return_statement', 'break_statement', 'continue_statement']:
                    insert_pos = child.start_byte  # Insert before control statement
                    break
                elif child.type not in ['comment', '}', '{', ';']:
                    break

            if insert_pos is None:
                insert_pos = children[-1].start_byte  # before closing }

            marker_positions.append(insert_pos)

        for child in node.children:
            collect_marker_insertions(child, marker_positions)

    marker_positions = []
    collect_marker_insertions(tree.root_node, marker_positions)

    # === Generate marker definitions (formatted with printf(" b%db ", i)) ===
    marker_defs = b'#include <stdio.h>\n'
    marker_defs += b'static int idx = 1;\n'
    marker_defs += b'// === Marker Definitions ===\n'
    for i in range(len(marker_positions)):
        def_line = f'void  __attribute__ ((noinline)) marker_{i}(void) {{ printf(" b%db ", {i}); idx++; idx=0; }}\n'
        marker_defs += def_line.encode()
    marker_defs += b'\n'

    # === Insert marker_i(); calls ===
    modified_code = bytearray(code)
    offset = 0
    for i, pos in enumerate(sorted(marker_positions)):
        marker_call = f'\n    marker_{i}();'.encode()
        modified_code[pos + offset:pos + offset] = marker_call
        offset += len(marker_call)

    # === Prepend marker definitions at top of file ===
    final_code = marker_defs + modified_code

    with open(output_file, "wb") as f:
        f.write(final_code)

    print(f"âœ… Inserted {len(marker_positions)} marker calls and moved only function definitions before main()")


def parse_structural_map(c_path, key_list):
    # === Compile grammar and set up parser ===
    Language.build_library(
        'build/my-languages.so',
        ['/home/xxx/disk-dut/research/github/tree-sitter-c']
    )
    C_LANGUAGE = Language('build/my-languages.so', 'c')
    parser = Parser()
    parser.set_language(C_LANGUAGE)

    with open(c_path, 'r') as f:
        lines = f.readlines()
    code = "".join(lines)
    format_line = lines[-1].strip()
    format_codes = re.findall(r"\d{2}", format_line)

    tree = parser.parse(code.encode("utf8"))
    root_node = tree.root_node

    # === Locate main() function ===
    main_func = None
    for child in root_node.children:
        if child.type == "function_definition":
            decl = child.child_by_field_name("declarator")
            ident = decl.child_by_field_name("declarator")
            if ident and code[ident.start_byte:ident.end_byte] == "main":
                main_func = child
                break
    if not main_func:
        raise ValueError("main() not found")

    # === Collect strtol assignments inside main() ===
    assignments = []
    var_list = []

    def walk(node):
        if node.type == "expression_statement":
            expr = node.children[0]
            if expr.type == "assignment_expression":
                lhs = expr.child_by_field_name("left")
                rhs = expr.child_by_field_name("right")
                if rhs and rhs.type == "call_expression":
                    func = rhs.child_by_field_name("function")
                    if func and code[func.start_byte:func.end_byte] == "strtol":
                        var_name = code[lhs.start_byte:lhs.end_byte]
                        var_list.append(var_name)
                        assignments.append({
                            "var": var_name,
                            "start_point": node.start_point,
                            "end_point": node.end_point
                        })
        for child in node.children:
            walk(child)

    walk(main_func)

    # === Build var -> format map
    var_format_map = OrderedDict()
    for i, var in enumerate(var_list):
        if i < len(format_codes):
            var_format_map[var] = format_codes[i]

    # === Filter map
    filtered_map = OrderedDict((k, v) for k, v in var_format_map.items() if k in key_list)

    # === Remove excluded strtol lines from main()
    new_lines = lines.copy()
    for assign in assignments:
        var = assign["var"]
        if var not in filtered_map:
            start_row = assign["start_point"][0]
            new_lines[start_row] = ""  # remove this line only

    # === Replace last line with filtered format codes
    new_lines[-1] = "//" + "".join(filtered_map.values()) + "\n"

    # === Write back to the same file
    with open(c_path, "w") as f:
        f.writelines(new_lines)

    #print(f"âœ… Preserved formatting, removed {len(assignments) - len(filtered_map)} assignment(s), updated last line.")
    return filtered_map


def getBinaries(max_funcs, max_expr_complexity, csmith_options, max_sym_var, seed, cmd_list, opt):
    os.system("csmith-marker --max-funcs {} --max-expr-complexity {} \
     > testcase_1.c".format(max_funcs, max_expr_complexity, csmith_options))

    instrument_c_file("testcase_1.c", "testcase_1.c")
    # os.system("head test.c")
    if opt == False:
        print("[-] Do not select interesting variables to be symbolized ...")
        os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")
        os.system("grep if testcase_1.c > if-stmts.txt")
        # os.system("python3 filter.py")  # TODO enable filtering of loop
        os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")

        # only keep interesting but not too much symbolic variables as targets
        os.system("grep strtol testcase_1.c | wc -l > sym_count.txt")
        sym_var_count = subprocess.run(['cat', 'sym_count.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
        print("number of symbolic variables : ", sym_var_count)
        while int(sym_var_count) > max_sym_var or int(sym_var_count) == 0:
            print("Not an interesting test program, Re-gerating it for now ...")
            os.system("csmith-marker --max-funcs {} --max-expr-complexity {} {} \
                    > testcase_1.c".format(max_funcs, max_expr_complexity, csmith_options))
            # Filter useless symbolic variables
            os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")
            os.system("grep if testcase_1.c > if-stmts.txt")
            os.system("grep strtol testcase_1.c | wc -l > sym_count.txt")
            # os.system("python3 filter.py")  # enable filtering of loop
            os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")
            sym_var_count = subprocess.run(['cat', 'sym_count.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
            print("number of symbolic variables in while loop: ", sym_var_count)
            instrument_c_file("testcase_1.c", "testcase_1.c")

        # generate binaries
        index = 1
        binaries = list()
        for cmd in cmd_list:
            os.system(cmd + " testcase_1.c -o test{}".format(index))
            binaries.append("test{}".format(index))
            index += 1
    else:
        # Filter useless symbolic variables
        print("[+] OPT 1 is Applied: Do select interesting variables to be symbolized ...")
        os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")
        # only keep interesting but not too much symbolic variables as targets
        os.system("grep strtol testcase_1.c | wc -l > sym_count.txt")
        sym_var_count = subprocess.run(['cat', 'sym_count.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
        print("number of symbolic variables : ", sym_var_count)
        while int(sym_var_count) > max_sym_var or int(sym_var_count) == 0:
            print("Not an interesting test program, Re-gerating it for now ...")
            os.system("csmith-marker --max-funcs {} --max-expr-complexity {} {} \
                    > testcase_1.c".format(max_funcs, max_expr_complexity, csmith_options))
            # Filter useless symbolic variables
            os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")
            os.system("grep if testcase_1.c > if-stmts.txt")
            os.system("grep strtol testcase_1.c | wc -l > sym_count.txt")
            # os.system("python3 filter.py")  # enable filtering of loop
            os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")
            sym_var_count = subprocess.run(['cat', 'sym_count.txt'], stdout=subprocess.PIPE).stdout.decode("utf-8")
            print("number of symbolic variables in while loop: ", sym_var_count)
            instrument_c_file("testcase_1.c", "testcase_1.c")
        os.system("cp testcase_1.c testcases/")
        testcase_files = get_testcase_files()
        process_testcase_for_var_selection(testcase_files[0])
        with open("./results/testcase_1_results.json", "r") as f:
            var_data = json.load(f)
        if var_data == None:
            print("something wrong with the variable selection optimziation")
            exit(1)
        # print("var_data : ", var_data)
        if isinstance(var_data, dict):
            keys = var_data.keys()
            # print("Top-level keys:", list(keys))

        filtered_map = parse_structural_map('./testcase_1.c', keys)
        os.system("grep strto testcase_1.c | awk \'{print $1}\' > names.txt")

        index = 1
        binaries = list()
        for cmd in cmd_list:
            os.system(cmd + " testcase_1.c -o test{}".format(index))
            binaries.append("test{}".format(index))
            index += 1
    os.system("grep Seed testcase_1.c")
    return binaries


def exists_path_without_other_markers(entry, target, marker_addrs, graph):
    """Check if there exists a path from `entry` to `target` without passing through any other marker block."""
    visited = set()
    stack = [entry]

    while stack:
        node = stack.pop()
        if node == target:
            return True
        if node in visited or (node in marker_addrs and node != target):
            continue
        visited.add(node)
        stack.extend(graph.successors(node))

    return False


def reduce_marker_targets(binary_path, marker_index_str_set, dot_path="marker_reachability.dot", debug=False):
    import angr
    import networkx as nx
    from capstone import Cs, CS_ARCH_X86, CS_MODE_64
    from networkx.drawing.nx_pydot import write_dot

    def exists_path_without_other_markers(entry, target, marker_addrs, graph):
        visited = set()
        stack = [entry]
        while stack:
            node = stack.pop()
            if node == target:
                return True
            if node in visited or (node in marker_addrs and node != target):
                continue
            visited.add(node)
            stack.extend(graph.successors(node))
        return False

    proj = angr.Project(binary_path, auto_load_libs=False)
    cfg = proj.analyses.CFGEmulated(keep_state=True, context_sensitivity_level=1)

    # Step 1: Find marker functions
    marker_names = {f"marker_{idx}" for idx in marker_index_str_set}
    addr_to_idx = {f.addr: f.name.split("_")[1] for f in proj.kb.functions.values() if f.name in marker_names}

    # Step 2: Find basic blocks that call marker functions
    block_to_marker = {}
    cs = Cs(CS_ARCH_X86, CS_MODE_64)
    for node in cfg.graph.nodes:
        if not node.addr or not node.block:
            continue
        try:
            block_bytes = proj.loader.memory.load(node.addr, node.size)
        except Exception:
            continue
        for ins in cs.disasm(block_bytes, node.addr):
            if ins.mnemonic.startswith("call") and ins.op_str.startswith("0x"):
                try:
                    target = int(ins.op_str, 16)
                    if target in addr_to_idx:
                        block_to_marker[node.addr] = addr_to_idx[target]
                        break
                except:
                    continue

    addr_to_node = {n.addr: n for n in cfg.graph.nodes if n.addr is not None}
    marker_addrs = list(block_to_marker.keys())

    if len(marker_addrs) == 0:
        print("âš ï¸ No marker function address, return all for safety.")
        return None

    # Step 3: Build CFG and compute dominators
    main_func = proj.kb.functions.function(name="main")
    main_entry_addr = next(iter(main_func.blocks)).addr

    G = nx.DiGraph()
    for src in cfg.graph.nodes:
        if src.addr:
            for succ in cfg.graph.successors(src):
                if succ.addr:
                    G.add_edge(src.addr, succ.addr)

    if not nx.is_directed_acyclic_graph(G):
        G = nx.DiGraph(nx.dfs_tree(G, source=main_entry_addr))

    idoms = nx.immediate_dominators(G, main_entry_addr)

    # Step 4: Dominance-based marker reduction with additional path check
    kept_markers = set()
    marker_addr_set = set(marker_addrs)

    for marker_addr in marker_addrs:
        dominated = False
        current = marker_addr
        while current != main_entry_addr:
            dom = idoms.get(current, None)
            if dom is None:
                if not exists_path_without_other_markers(main_entry_addr, marker_addr, marker_addr_set, G):
                    dominated = True
                break
            if dom in block_to_marker and dom != marker_addr:
                dominated = True
                break
            current = dom

        if not dominated:
            kept_markers.add(block_to_marker[marker_addr])

    # Step 5: Visualization
    if debug:
        marker_graph = nx.DiGraph()
        for addr in marker_addrs:
            idx = block_to_marker[addr]
            marker_graph.add_node(f"marker_{idx}", label=f"marker_{idx}\\n0x{addr:x}")
        write_dot(marker_graph, dot_path)
        print(f"âœ… Marker CFG nodes written to: {dot_path}")
        print(f"âœ… Original markers: {sorted(marker_index_str_set, key=lambda x: int(x))}")
        print(f"âœ… Reduced markers: {sorted(kept_markers, key=lambda x: int(x))}")

    return sorted(kept_markers, key=lambda x: int(x))


def generate_args_from_format(fmt: str):
    """
    Generate a list of hex string arguments from a format string like "010414".
    Each pair means (sign, size): 0=signed, 1=unsigned; size in bytes: 1, 2, 4, 8.
    Arguments are returned in hex format (e.g., '7F') suitable for strtol(..., 16).
    """
    args = []
    if len(fmt) % 2 != 0:
        raise ValueError("Format string must have even number of digits")
    for i in range(0, len(fmt), 2):
        sign = fmt[i]
        size = int(fmt[i + 1])
        if size not in [1, 2, 4, 8]:
            raise ValueError(f"Invalid size: {size}")
        bits = size * 8
        if sign == '0':  # signed
            val = random.randint(-(2 ** (bits - 1)), 2 ** (bits - 1) - 1)
            val &= (2 ** bits - 1)  # wrap negative values like 2's complement
        elif sign == '1':  # unsigned
            val = random.randint(0, 2 ** bits - 1)
        else:
            raise ValueError(f"Invalid sign: {sign}")
        args.append(format(val, 'X'))  # upper-case hex string without "0x"
    return args


def run_binary(binary_path, args):
    """
    Run the binary with given arguments and return the output.
    Logs the full command and output in one line.
    Detects if a segmentation fault occurred.
    """
    # Save input arguments to temp.txt
    with open("temp.txt", "w") as f:
        f.write(' '.join(args))

    try:
        result = subprocess.run(
            [binary_path] + args,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            timeout=5
        )
        output = result.stdout.strip().replace("\n", " ")  # flatten output
        cmd_str = ' '.join([binary_path] + args)

        if result.returncode < 0:
            print(f"CMD: {cmd_str} | OUT: [SEGMENTATION FAULT]")
            num = int(random.random()*100000000)
            os.system("cp testcase_1.c wrong-{}-candidate-crash.c".format(num))
            os.system("cat temp.txt >> wrong-{}-candidate-crash.c".format(num))
            return "[SEGMENTATION FAULT]"

        # print(f"CMD: {cmd_str} | OUT: {output}")
        return output

    except subprocess.TimeoutExpired:
        #print(f"CMD: {binary_path} {' '.join(args)} | OUT: [TIMEOUT]")
        return "[TIMEOUT]"
    except Exception as e:
        #print(f"CMD: {binary_path} {' '.join(args)} | OUT: [ERROR: {e}]")
        return f"[ERROR: {e}]"


def generate_and_check_until_found(fmt: str, binary_path: str, marker: str, timeout=60):
    """
    Repeatedly generate arguments and run binary until the marker is found or timeout.
    Returns (success, elapsed_time, final_args)
    """
    start = time.time()
    tries = 0
    while time.time() - start < timeout:
        args = generate_args_from_format(fmt)
        output = run_binary(binary_path, args)
        tries += 1
        if output and marker in output:
            elapsed = time.time() - start
            print(f"âœ… Marker '{marker}' found after {elapsed:.2f}s in try {tries}")
            return True, elapsed, args
    print(f"â° Timeout reached after {timeout}s, marker '{marker}' not found.")
    return False, timeout, []


def testMain(binary, is_reduced, perform_se):
    # get all markers
    # TODO Step 1: find interesting markers
    global count_se, count_markers_org, count_markers_reduced
    global count_max_mem, time_duration, bool_diff, count_diff
    markers = getInterestingMarkerSetV5()
    print("marker interesting : ", markers)
   
    # check markers first, if it is empty, just return
    if len(markers["test1"]) == 0 and len(markers["test2"]) == 0:
        print("No interesting markers ...")
        bool_diff = 0
        return
    count_diff += 1
    os.system("grep 'Seed' testcase_1.c | awk '{print $3}' >> seeds_interesting.txt")
    for b in binary:
        bool_diff = 1
        print("Dealing with ", b)
        
        if len(markers[b]) == 0:
            print("not interesting binary, skip it ...")
            continue
        
        if is_reduced == True:
            # reduced_markers = filter_intersection(markers[b], "{}.dot".format(b), b)
            if len(markers[b]) != 1:
                print("[+] OPT 2 is Applied: Use the optimization for marker reduction ... ")
                reduced_markers = reduce_marker_targets(b, markers[b])
                if reduced_markers == None:
                    print("skip this binary ...")
                    continue
            else:
                print("[+] OPT 2 is Omitted: since there is only one element in the set ... ")
                reduced_markers = markers[b]
        else:
            print("[-] Do not use the optimization for marker reduction")
            reduced_markers = markers[b]
        if reduced_markers is None:
           print("Do not exist in this CFG ...")
           continue
       
        count_markers_org += len(markers[b])
        count_markers_reduced += len(reduced_markers)
        if len(markers[b]) != len(reduced_markers):
           print("orig_markers : ", markers[b])
           print("reduced_markers : ", reduced_markers)
           #if len(reduced_markers) >= 2:
           # exit(1)
        # print("markers : ", markers)
        for inter in reduced_markers:
            count_se += 1

            if perform_se == True:
                print("do explore : ", inter)
                print("This is an interesting path, starting to conducting binary symbolic execution")
                
                # Set timeout duration
                se_found = False
                TIMEOUT = 60
                q = Queue()
                p = Process(target=wrapper, args=(b, inter, q))

                p.start()
                start_time = time.time()

                max_mem = 0

                # Polling loop with timeout check
                while p.is_alive():
                    elapsed = time.time() - start_time
                    if elapsed > TIMEOUT:
                        print(f"â° Timeout reached after {elapsed:.1f} seconds. Terminating process.")
                        p.terminate()
                        p.join()  # Ensure process is cleaned up
                        break

                    mem = measure_memory(p.pid)
                    max_mem = max(max_mem, mem)
                    time.sleep(0.1)

                # Check if process finished before timeout
                if not p.is_alive():
                    p.join()
                    if not q.empty():
                        out = q.get()
                        if out == True:
                            se_found = True
                        print("Function returned:", out)
                        # print("Captured stdout:")
                        # print(out['stdout'])
                    else:
                        print("Process ended without returning anything.")
                end_time = time.time()
                duration = end_time - start_time
                print(f"â±ï¸ Duration: {duration:.2f} seconds")
                print(f"ðŸ” Peak memory usage: {max_mem:.2f} MB")
                count_max_mem = max_mem
                time_duration = duration

                success = False
                elapsed_fuzz = -1
                if se_found == False:
                    print("+++ Start Fuzzing (as SE can not find the target) ...")
                    results = subprocess.run(['tail', '-1', 'testcase_1.c'], stdout=subprocess.PIPE).stdout.decode("utf-8")
                    results = results[2:-1]
                    # print("results : ", results)
                    binary = "./"+ b
                    inter = 'b' + inter + 'b'
                    print("marker for fuzzing : ", inter)
                    success, elapsed_fuzz, final_args = generate_and_check_until_found(results, binary, inter, timeout=60)
                    print("+++ Fuzzing done. fuzzing time = ", elapsed_fuzz)
                    # exit(1)
                if success == True:
                    num = int(random.random()*100000000)
                    os.system("cp testcase_1.c wrong-{}-candidate-fuzz-found.c".format(num))
                    os.system("cat temp.txt >> wrong-{}-candidate-fuzz-found.c".format(num))

                # âœ… Append result to log file
                excel_path = "log_se_execution_results.xlsx"
                if os.path.exists(excel_path):
                    wb = load_workbook(excel_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["Run", "Marker ID", "Duration_se (s)", "succ_se", "Duration_fuzz (s)", "succ_fuzz", "Max Memory (MB)"])  # Header
                ws.append([count_se, inter, round(duration, 2), int(se_found), round(elapsed_fuzz, 2), int(success), round(max_mem, 2)])
                wb.save(excel_path)

def measure_memory(pid):
    try:
        process = psutil.Process(pid)
        mem_mb = process.memory_info().rss / (1024 * 1024)  # in MB
        return mem_mb
    except psutil.NoSuchProcess:
        return 0

def log_to_excel(log_path, i, count_var_org, count_var_reduced, bool_diff,
                 count_markers_org, count_markers_reduced, count_se):
    if os.path.exists(log_path):
        wb = load_workbook(log_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Test No", "count_var_org", "count_var_reduced", "bool_diff",
            "count_markers_org", "count_markers_reduced", "count_se"
        ])

    ws.append([
        i, count_var_org, count_var_reduced, count_diff,
        count_markers_org, count_markers_reduced, count_se
    ])
    wb.save(log_path)

TIMEOUT_SECONDS = 864000  # Set running timeout
count_se = 0
count_markers_org = 0
count_markers_reduced = 0
count_var_org = 0
count_var_reduced = 0
count_found = 0
count_max_mem = 0
time_duration = 0
bool_diff = 0
count_diff = 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-var-selection", action="store_false", dest="var_selection",
                    help="Disable variable selection optimization in getBinaries (default: enabled)")
    parser.add_argument("--no-marker-reduction", action="store_false", dest="marker_reduction",
                    help="Disable marker reduction optimization in testMain (default: enabled)")
    parser.add_argument("--no-se", action="store_false", dest="perform_se",
                    help="Disable symbolic execution (default: disabled)")
    parser.add_argument("--num-tests", type=int, default=100,
                    help="Number of test iterations to run (default: 100000)")
    parser.set_defaults(var_selection=True, marker_reduction=True, perform_se=True)
    args = parser.parse_args()

    print("### Test main ###")

    cmd_list = []
    with open("./compiler_test.in") as file_in:
        for line in file_in:
            cmd_list.append(line.rstrip('\n'))
    print("cmd_list : ", cmd_list)

    start_time = time.time()

    s = 0
    for i in range(args.num_tests):
        elapsed = time.time() - start_time
        print("Execution time: ", elapsed)
        if elapsed > TIMEOUT_SECONDS:
            print(f"â° Timeout reached after {elapsed:.1f} seconds. Exiting loop.")
            exit(1)
        print("\nGeting binaries ... ", i)
        binaries = getBinaries(10, 8, "", 200, s, cmd_list, args.var_selection)
        print("Analyzing binaries ...")
        testMain(binaries, args.marker_reduction, args.perform_se)
        os.system("rm -rf ./testdbs/*")
        print(f"############### No.{i} count_var_org = {count_var_org}, count_var_reduced = {count_var_reduced}; count_diff = {count_diff}; count_org_marker = {count_markers_org}; count_reduced_marker = {count_markers_reduced}; count_se = {count_se}")
    log_to_excel(
    "log_overall_results.xlsx", i+1,
    count_var_org, count_var_reduced, count_diff,
    count_markers_org, count_markers_reduced, count_se
    )

